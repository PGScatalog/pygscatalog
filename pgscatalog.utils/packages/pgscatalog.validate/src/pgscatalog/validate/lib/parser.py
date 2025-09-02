import abc
import csv
import gzip
from abc import abstractmethod
from collections.abc import Iterable
from pathlib import Path
from typing import List, TextIO, Generator

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

import pgscatalog.validate.lib.errors as errors


def skip_comments(file: TextIO) -> Generator:
    """Reads the input file and returns each line as a generator, skipping the commented lines starting with #."""
    return (line for line in file if not line.startswith("#"))


class ScoreFileParser(abc.ABC):
    """Abstract file parser. Implement it to add another supported file format."""

    file_path: Path

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        if not self.file_path.is_file():
            raise FileNotFoundError(f"Error: File not found: {self.file_path}")

    @abstractmethod
    def get_column_names(self) -> List[str]:
        raise NotImplementedError("Not implemented")

    @abstractmethod
    def get_variants(self) -> Iterable:
        raise NotImplementedError("Not implemented")


class TSVFileParser(ScoreFileParser):

    # Abstracted open function, in case the file is gzipped
    open_fn = open

    def __init__(self, file_path: str):
        super().__init__(file_path)
        if file_path.endswith(".gz"):
            self.open_fn = gzip.open

    def get_column_names(self) -> List[str]:
        with self.open_fn(self.file_path, mode="rt", encoding="utf-8", errors="replace") as file:
            reader = csv.reader(skip_comments(file), delimiter="\t")
            return next(reader)  # Get the first row (header)

    def get_variants(self) -> Iterable:
        with self.open_fn(self.file_path, mode="rt", encoding="utf-8", errors="replace") as file:
            yield from csv.DictReader(skip_comments(file), delimiter="\t")


class SpreadsheetFileParser(ScoreFileParser):
    """Spreadsheet file parser.
    The libreoffice format is causing issues with data types at the moment. Prefer the .xlsx format."""

    worksheet: Worksheet

    def __init__(self, file_path: str):
        super().__init__(file_path)
        self.worksheet = load_workbook(self.file_path, read_only=True).worksheets[0]

    def get_column_names(self) -> List[str]:
        row = next(self.worksheet.iter_rows(values_only=True))
        return list(row)

    def get_variants(self) -> Iterable:
        rows = self.worksheet.iter_rows(values_only=True)
        header = next(rows)

        for row in rows:
            yield dict(zip(header, row))  # Map headers to row values

    @staticmethod
    def _str_to_bool(value: str) -> bool | None:
        if value:
            if value.lower() in ("yes", "true", "true()", "=true()", "1"):
                return True
            elif value.lower() in ("no", "false", "false()", "=false()", "0"):
                return False
            else:
                raise ValueError(errors.NOT_A_BOOLEAN)
        else:
            return None
